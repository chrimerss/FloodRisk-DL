#!/usr/bin/env python
# coding: utf-8

# Import necessary libraries
import os
import sys
import numpy as np
import matplotlib.pyplot as plt
import torch
import torch.nn as nn
from torch.utils.data import DataLoader
import torchvision.transforms as transforms
import json
import rasterio
from rasterio.windows import Window
from glob import glob
import cartopy.crs as ccrs
import cartopy.feature as cfeature
import matplotlib.colors as mcolors
import xarray as xr
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

# Add project root to path for custom modules
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Import custom modules
from src.data.flood_data_module import FloodDataModule
from src.models.flood_model import FloodPredictionModel

# Define data directories
DATA_DIR = "/home/users/li1995/global_flood/UrbanFloods2D/dataset"
FLOOD_SAMPLE_DIR = "/home/users/li1995/global_flood/UrbanFloods2D/sample"
OUTPUT_DIR = "output"

# Create output directory if it doesn't exist
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Load config
from hydra import compose, initialize
from omegaconf import OmegaConf

# Load config
config = OmegaConf.load('/home/users/li1995/global_flood/FloodRisk-DL/configs/model_config.yaml')

# Set model parameters
config['model']['window_size'] = 8
config['model']['num_layers'] = 4

# Load normalization statistics
with open('../src/normalization_stats.json', 'r') as f:
    norm_stats = json.load(f)

# Define a function to read and process geotiff data
def read_geotiff(file_path):
    """Read a GeoTIFF file and return data and metadata."""
    with rasterio.open(file_path) as src:
        data = src.read(1)  # Read first band
        metadata = {
            'transform': src.transform,
            'crs': src.crs,
            'bounds': src.bounds,
            'height': src.height,
            'width': src.width
        }
    return data, metadata

# Define a function to crop data from geotiff
def crop_random_tile(dem_data, dem_meta, flood_data, flood_meta, crop_size=1024, margin=10):
    """Crop a random tile from the input data."""
    height, width = dem_data.shape
    
    # Choose random coordinates with margin
    i = np.random.randint(margin, height - crop_size - margin)
    j = np.random.randint(margin, width - crop_size - margin)
    
    # Crop data
    dem_crop = dem_data[i:i+crop_size, j:j+crop_size]
    flood_crop = flood_data[i:i+crop_size, j:j+crop_size]
    
    # Update transform for the cropped area
    window = Window(j, i, crop_size, crop_size)
    dem_transform = rasterio.windows.transform(window, dem_meta['transform'])
    flood_transform = rasterio.windows.transform(window, flood_meta['transform'])
    
    # Create cropped metadata
    dem_crop_meta = dem_meta.copy()
    dem_crop_meta['transform'] = dem_transform
    dem_crop_meta['height'] = crop_size
    dem_crop_meta['width'] = crop_size
    dem_crop_meta['crop_window'] = (i, j, crop_size, crop_size)
    
    flood_crop_meta = flood_meta.copy()
    flood_crop_meta['transform'] = flood_transform
    flood_crop_meta['height'] = crop_size
    flood_crop_meta['width'] = crop_size
    
    return dem_crop, dem_crop_meta, flood_crop, flood_crop_meta

# Define a function to normalize data
def normalize_data(dem, depth, rainfall, city_id):
    """Normalize input data based on statistics."""
    stats = norm_stats[city_id]
    
    # Normalize DEM
    dem_normalized = (dem - stats['dem_mean']) / (stats['dem_std'] + 1e-8)
    
    # Normalize flood depth (no mean subtraction for depth)
    depth_normalized = depth / (stats['depth_std'] + 1e-8)
    
    # No normalization for rainfall
    
    return dem_normalized, depth_normalized, rainfall

# Define a function to denormalize data
def denormalize_data(depth_normalized, city_id):
    """Denormalize flood depth prediction."""
    stats = norm_stats[city_id]
    
    # Denormalize flood depth
    depth = depth_normalized * (stats['depth_std'] + 1e-8)
    
    return depth

# Define a function to prepare model input
def prepare_model_input(dem_normalized, rainfall, device='cuda'):
    """Prepare input tensor for the model."""
    # Create rainfall channel (same value across the entire image)
    rainfall_channel = np.ones_like(dem_normalized) * rainfall
    
    # Stack DEM and rainfall channels
    model_input = np.stack([dem_normalized, rainfall_channel], axis=0)
    
    # Convert to torch tensor
    model_input = torch.from_numpy(model_input).float().unsqueeze(0).to(device)
    
    return model_input

# Define a function to visualize results
def visualize_results(dem, target_depth, predicted_depth, metadata, city_id, rainfall, output_path=None):
    """Visualize input DEM, target and predicted flood depths using cartopy."""
    # Get bounds in lat/lon format
    bounds = metadata['bounds']
    transform = metadata['transform']
    crs = metadata['crs']
    
    # Create a figure with four subplots (added comparison column)
    fig, axes = plt.subplots(1, 4, figsize=(24, 6), subplot_kw={'projection': ccrs.PlateCarree()}, dpi=350)
    
    # Common plot settings
    for ax in axes:
        ax.add_feature(cfeature.LAND)
        ax.add_feature(cfeature.OCEAN)
        ax.add_feature(cfeature.COASTLINE)
        ax.add_feature(cfeature.BORDERS, linestyle=':')
        ax.add_feature(cfeature.LAKES, alpha=0.5)
        ax.add_feature(cfeature.RIVERS)
    
    # Plot DEM
    dem_plot = axes[0].imshow(
        dem, 
        cmap='terrain', 
        extent=[bounds.left, bounds.right, bounds.bottom, bounds.top],
        transform=ccrs.PlateCarree()
    )
    axes[0].set_title(f'DEM - {city_id}')
    plt.colorbar(dem_plot, ax=axes[0], label='Elevation (m)', fraction=0.046)
    
    # Define common colormap for depth plots
    max_depth = max(np.max(target_depth), np.max(predicted_depth))
    norm = mcolors.Normalize(vmin=0, vmax=max_depth)
    cmap = plt.cm.Blues
    
    # Plot target flood depth
    target_plot = axes[1].imshow(
        target_depth, 
        cmap=cmap, 
        norm=norm,
        extent=[bounds.left, bounds.right, bounds.bottom, bounds.top],
        transform=ccrs.PlateCarree()
    )
    axes[1].set_title(f'Target Flood Depth - {rainfall}mm')
    plt.colorbar(target_plot, ax=axes[1], label='Depth (m)', fraction=0.046)
    
    # Plot predicted flood depth
    pred_plot = axes[2].imshow(
        predicted_depth, 
        cmap=cmap, 
        norm=norm,
        extent=[bounds.left, bounds.right, bounds.bottom, bounds.top],
        transform=ccrs.PlateCarree()
    )
    axes[2].set_title('Predicted Flood Depth')
    plt.colorbar(pred_plot, ax=axes[2], label='Depth (m)', fraction=0.046)
    
    # Create comparison mask
    # Define thresholds for binary flood/no-flood classification (>0.05m is flood)
    threshold = 0.05
    target_binary = target_depth > threshold
    pred_binary = predicted_depth > threshold
    
    # Create categorical comparison
    # 0: both no flood (gray)
    # 1: hit - both predict flood (blue)
    # 2: miss - target has flood but model misses (green)
    # 3: false alarm - model predicts flood but target doesn't have it (red)
    comparison = np.zeros_like(target_depth, dtype=np.uint8)
    comparison[~target_binary & ~pred_binary] = 0  # Both no flood (gray)
    comparison[target_binary & pred_binary] = 1    # Hit (blue)
    comparison[target_binary & ~pred_binary] = 2   # Miss (green)
    comparison[~target_binary & pred_binary] = 3   # False alarm (red)
    
    # Custom colormap for comparison
    comparison_colors = ['gray', 'blue', 'green', 'red']
    comparison_cmap = mcolors.ListedColormap(comparison_colors)
    
    # Plot comparison
    comparison_plot = axes[3].imshow(
        comparison,
        cmap=comparison_cmap,
        vmin=0,
        vmax=3,
        extent=[bounds.left, bounds.right, bounds.bottom, bounds.top],
        transform=ccrs.PlateCarree()
    )
    axes[3].set_title('Comparison')
    
    # Create custom legend for comparison
    from matplotlib.patches import Patch
    legend_elements = [
        Patch(facecolor='gray', label='Both no flood'),
        Patch(facecolor='blue', label='Hit (both predict flood)'),
        Patch(facecolor='green', label='Miss (model misses flood)'),
        Patch(facecolor='red', label='False alarm')
    ]
    axes[3].legend(handles=legend_elements, loc='lower right')
    
    plt.tight_layout()
    
    # Save figure if output path is provided
    if output_path:
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        print(f"Figure saved to {output_path}")
    
    plt.close()

# Calculate categorical metrics for the flood prediction
def calculate_categorical_metrics(target, prediction, threshold=0.05):
    """
    Calculate categorical verification metrics for flood prediction.
    
    Args:
        target: Target flood depth array
        prediction: Predicted flood depth array
        threshold: Threshold for flood/no-flood classification (default: 0.05m)
        
    Returns:
        Dictionary of categorical metrics
    """
    # Create binary masks
    target_binary = target > threshold
    pred_binary = prediction > threshold
    
    # Calculate contingency table values
    hits = np.sum((target_binary) & (pred_binary))  # Both predicted and occurred
    misses = np.sum((target_binary) & (~pred_binary))  # Occurred but not predicted
    false_alarms = np.sum((~target_binary) & (pred_binary))  # Predicted but didn't occur
    correct_negatives = np.sum((~target_binary) & (~pred_binary))  # Neither predicted nor occurred
    
    # Calculate metrics
    total = hits + misses + false_alarms + correct_negatives
    
    # Proportion correct (accuracy)
    accuracy = (hits + correct_negatives) / total if total > 0 else 0
    
    # Hit rate (probability of detection)
    hit_rate = hits / (hits + misses) if (hits + misses) > 0 else 0
    
    # False alarm rate
    false_alarm_rate = false_alarms / (false_alarms + correct_negatives) if (false_alarms + correct_negatives) > 0 else 0
    
    # Critical success index (threat score)
    csi = hits / (hits + misses + false_alarms) if (hits + misses + false_alarms) > 0 else 0
    
    # Bias score
    bias = (hits + false_alarms) / (hits + misses) if (hits + misses) > 0 else 0
    
    # F1 score
    precision = hits / (hits + false_alarms) if (hits + false_alarms) > 0 else 0
    recall = hit_rate  # Same as POD
    f1 = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0
    
    return {
        'Accuracy': accuracy,
        'Hit Rate (POD)': hit_rate,
        'False Alarm Rate': false_alarm_rate,
        'Critical Success Index': csi,
        'Bias Score': bias,
        'Precision': precision,
        'F1 Score': f1,
        'Contingency Table': {
            'Hits': int(hits),
            'Misses': int(misses),
            'False Alarms': int(false_alarms),
            'Correct Negatives': int(correct_negatives)
        }
    }

# Modify the process_city function to include categorical metrics
def process_city(city_id, rainfall_level, model, device='cuda'):
    """Process a single city with the given rainfall level."""
    print(f"Processing {city_id} with {rainfall_level}mm rainfall")
    
    # Load data paths
    dem_file = os.path.join(DATA_DIR, city_id, f"{city_id}_DEM.tif")
    flood_file = os.path.join(FLOOD_SAMPLE_DIR, f"{city_id}_{rainfall_level}_max.tif")
    
    # Read data
    dem_data, dem_meta = read_geotiff(dem_file)
    flood_data, flood_meta = read_geotiff(flood_file)
    
    # Crop random tile
    dem_crop, dem_crop_meta, flood_crop, flood_crop_meta = crop_random_tile(
        dem_data, dem_meta, flood_data, flood_meta
    )
    
    # Extract rainfall value as integer (remove 'mm')
    rainfall = int(rainfall_level.replace('mm', ''))
    
    # Normalize data
    dem_normalized, depth_normalized, _ = normalize_data(dem_crop, flood_crop, rainfall, city_id)
    
    # Prepare model input
    model_input = prepare_model_input(dem_normalized, rainfall, device)
    
    # Get model prediction
    model.eval()
    with torch.no_grad():
        prediction = model(model_input)
    
    # Convert prediction to numpy
    prediction_np = prediction.cpu().numpy().squeeze()
    
    # Denormalize prediction
    prediction_denorm = denormalize_data(prediction_np, city_id)
    
    # Calculate categorical metrics
    categorical_metrics = calculate_categorical_metrics(flood_crop, prediction_denorm)
    
    # Print categorical metrics
    print(f"\nCategorical Metrics for {city_id}:")
    for metric, value in categorical_metrics.items():
        if metric != 'Contingency Table':
            print(f"  {metric}: {value:.4f}")
    
    print("  Contingency Table:")
    for key, value in categorical_metrics['Contingency Table'].items():
        print(f"    {key}: {value}")
    
    # Visualize results
    output_path = os.path.join(OUTPUT_DIR, f"{city_id}_{rainfall_level}_prediction.png")
    visualize_results(
        dem_crop, flood_crop, prediction_denorm, 
        dem_crop_meta, city_id, rainfall_level, 
        output_path
    )
    
    return dem_crop, flood_crop, prediction_denorm, dem_crop_meta, categorical_metrics

# Calculate evaluation metrics
def calculate_metrics(target, prediction):
    """Calculate evaluation metrics between target and prediction."""
    # Mean Squared Error
    mse = np.mean((target - prediction) ** 2)
    
    # Mean Absolute Error
    mae = np.mean(np.abs(target - prediction))
    
    # Root Mean Squared Error
    rmse = np.sqrt(mse)
    
    # Coefficient of determination (R²)
    ss_tot = np.sum((target - np.mean(target)) ** 2)
    ss_res = np.sum((target - prediction) ** 2)
    r2 = 1 - (ss_res / (ss_tot + 1e-10))
    
    return {
        'MSE': mse,
        'MAE': mae,
        'RMSE': rmse,
        'R²': r2
    }

# Function to create and save a comparative visualization of all three datasets
def create_comparison_plot(train_data, test_data, val_data, output_path=None):
    """Create a comparison plot of all three datasets."""
    fig, axes = plt.subplots(3, 4, figsize=(24, 18))  # Changed to 4 columns
    
    # Data tuples: (city_id, dem, target, pred, meta)
    datasets = [
        train_data,
        test_data,
        val_data
    ]
    
    # Row titles
    row_titles = ['Training', 'Testing', 'Validation']
    
    # Column titles
    col_titles = ['DEM', 'Target Flood Depth', 'Predicted Flood Depth', 'Comparison']
    
    # Set global max for depth colormap
    city_id, dem, target, pred, meta = train_data
    max_depth = max(np.max(target), np.max(pred))
    
    city_id, dem, target, pred, meta = test_data
    max_depth = max(max_depth, np.max(target), np.max(pred))
    
    city_id, dem, target, pred, meta = val_data
    max_depth = max(max_depth, np.max(target), np.max(pred))
    
    depth_norm = mcolors.Normalize(vmin=0, vmax=2)
    depth_cmap = plt.cm.Blues
    
    # Define threshold for binary flood/no-flood classification
    threshold = 0.05
    
    # Custom colormap for comparison
    comparison_colors = ['gray', 'blue', 'green', 'red']
    comparison_cmap = mcolors.ListedColormap(comparison_colors)
    
    # Legend elements for comparison plot
    from matplotlib.patches import Patch
    legend_elements = [
        Patch(facecolor='gray', label='Both no flood'),
        Patch(facecolor='blue', label='Hit (both predict flood)'),
        Patch(facecolor='green', label='Miss (model misses flood)'),
        Patch(facecolor='red', label='False alarm')
    ]
    
    # Plot each dataset
    for i, (city_id, dem, target, pred, meta) in enumerate(datasets):
        # Plot DEM
        dem_plot = axes[i, 0].imshow(dem, cmap='terrain')
        axes[i, 0].set_title(f'{col_titles[0]} - {city_id}')
        plt.colorbar(dem_plot, ax=axes[i, 0], label='Elevation (m)', fraction=0.046)
        
        # Plot target
        target_plot = axes[i, 1].imshow(target, cmap=depth_cmap, norm=depth_norm)
        axes[i, 1].set_title(f'{col_titles[1]}')
        plt.colorbar(target_plot, ax=axes[i, 1], label='Depth (m)', fraction=0.046)
        
        # Plot prediction
        pred_plot = axes[i, 2].imshow(pred, cmap=depth_cmap, norm=depth_norm)
        axes[i, 2].set_title(f'{col_titles[2]}')
        plt.colorbar(pred_plot, ax=axes[i, 2], label='Depth (m)', fraction=0.046)
        
        # Create binary masks for comparison
        target_binary = target > threshold
        pred_binary = pred > threshold
        
        # Create categorical comparison
        # 0: both no flood (gray)
        # 1: hit - both predict flood (blue)
        # 2: miss - target has flood but model misses (green)
        # 3: false alarm - model predicts flood but target doesn't have it (red)
        comparison = np.zeros_like(target, dtype=np.uint8)
        comparison[~target_binary & ~pred_binary] = 0  # Both no flood (gray)
        comparison[target_binary & pred_binary] = 1    # Hit (blue)
        comparison[target_binary & ~pred_binary] = 2   # Miss (green)
        comparison[~target_binary & pred_binary] = 3   # False alarm (red)
        
        # Plot comparison
        comp_plot = axes[i, 3].imshow(comparison, cmap=comparison_cmap, vmin=0, vmax=3)
        axes[i, 3].set_title(f'{col_titles[3]}')
        
        # Add legend to the first comparison plot only to avoid repetition
        if i == 0:
            axes[i, 3].legend(handles=legend_elements, loc='lower right')
        
        # Add row title
        fig.text(0.01, 0.75 - i*0.33, row_titles[i], va='center', ha='center', 
                 rotation='vertical', fontsize=14, fontweight='bold')
    
    plt.tight_layout()
    
    # Save figure if output path is provided
    if output_path:
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        print(f"Comparison figure saved to {output_path}")
    
    plt.close()

# Function to create and save visualizations of metrics
def create_metrics_visualizations(train_metrics, test_metrics, val_metrics, 
                               train_cat_metrics, test_cat_metrics, val_cat_metrics,
                               train_city, test_city, val_city, output_dir):
    """Create bar charts visualizing model performance metrics."""
    # Set up plot style
    plt.style.use('seaborn-v0_8-whitegrid')
    
    # Define datasets for labels
    datasets = [f"Training\n({train_city})", f"Testing\n({test_city})", f"Validation\n({val_city})"]
    
    # Create comparison plots (across domains)
    create_comparison_metric_plots(
        train_metrics, test_metrics, val_metrics,
        train_cat_metrics, test_cat_metrics, val_cat_metrics,
        train_city, test_city, val_city, output_dir
    )
    
    # Create domain-specific visualizations
    domains = [
        (train_city, train_metrics, train_cat_metrics, "Training"),
        (test_city, test_metrics, test_cat_metrics, "Testing"),
        (val_city, val_metrics, val_cat_metrics, "Validation")
    ]
    
    for city, reg_metrics, cat_metrics, dataset_type in domains:
        create_domain_metrics_plot(city, reg_metrics, cat_metrics, dataset_type, output_dir)


def create_domain_metrics_plot(city, reg_metrics, cat_metrics, dataset_type, output_dir):
    """Create a comprehensive metrics visualization for a single domain."""
    # Create a comprehensive figure
    fig, axes = plt.subplots(2, 2, figsize=(14, 12))
    fig.suptitle(f"{dataset_type} Domain: {city} - Performance Metrics", fontsize=16, fontweight='bold')
    
    # Plot regression metrics
    regression_metrics = ['MSE', 'MAE', 'RMSE', 'R²']
    reg_values = [reg_metrics[m] for m in regression_metrics]
    
    ax = axes[0, 0]
    bars = ax.bar(regression_metrics, reg_values, color='#1f77b4')
    ax.set_title(f"Regression Metrics", fontsize=12, fontweight='bold')
    ax.set_ylabel('Value')
    
    # Add value labels
    for bar in bars:
        height = bar.get_height()
        ax.annotate(f'{height:.4f}',
                   xy=(bar.get_x() + bar.get_width() / 2, height),
                   xytext=(0, 3),
                   textcoords="offset points",
                   ha='center', va='bottom', rotation=0, fontsize=10)
    
    # Plot categorical metrics
    categorical_metrics = ['Accuracy', 'Hit Rate (POD)', 'False Alarm Rate', 'F1 Score', 'Critical Success Index']
    cat_values = [cat_metrics[m] for m in categorical_metrics]
    
    ax = axes[0, 1]
    bars = ax.bar(categorical_metrics, cat_values, color='#ff7f0e')
    ax.set_title(f"Categorical Metrics", fontsize=12, fontweight='bold')
    ax.set_ylim(0, 1.0)
    plt.setp(ax.get_xticklabels(), rotation=20, ha='right')
    
    # Add value labels
    for bar in bars:
        height = bar.get_height()
        ax.annotate(f'{height:.4f}',
                   xy=(bar.get_x() + bar.get_width() / 2, height),
                   xytext=(0, 3),
                   textcoords="offset points",
                   ha='center', va='bottom', rotation=0, fontsize=10)
    
    # Plot contingency table as pie chart
    contingency_metrics = ['Hits', 'Misses', 'False Alarms', 'Correct Negatives']
    contingency_values = [cat_metrics['Contingency Table'][m] for m in contingency_metrics]
    
    # Calculate percentages
    total = sum(contingency_values)
    percentages = [val/total*100 for val in contingency_values]
    
    # Create pie chart of contingency table
    ax = axes[1, 0]
    colors = ['blue', 'green', 'red', 'gray']
    wedges, texts, autotexts = ax.pie(
        contingency_values, 
        labels=contingency_metrics,
        autopct='%1.1f%%',
        startangle=90,
        colors=colors
    )
    ax.set_title(f"Contingency Table Distribution", fontsize=12, fontweight='bold')
    
    # Make pie chart text more readable
    for text in texts:
        text.set_fontsize(9)
    for autotext in autotexts:
        autotext.set_fontsize(9)
        autotext.set_fontweight('bold')
        autotext.set_color('white')
    
    # Plot contingency values as bar chart
    ax = axes[1, 1]
    bars = ax.bar(contingency_metrics, contingency_values, color=colors)
    ax.set_title(f"Contingency Table Counts", fontsize=12, fontweight='bold')
    ax.set_ylabel('Count')
    
    # Add value labels
    for bar in bars:
        height = bar.get_height()
        ax.annotate(f'{int(height)}',
                   xy=(bar.get_x() + bar.get_width() / 2, height),
                   xytext=(0, 3),
                   textcoords="offset points",
                   ha='center', va='bottom', fontsize=10)
    
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, f"{city}_metrics.png"), dpi=300, bbox_inches='tight')
    plt.close()


def create_comparison_metric_plots(train_metrics, test_metrics, val_metrics, 
                               train_cat_metrics, test_cat_metrics, val_cat_metrics,
                               train_city, test_city, val_city, output_dir):
    """Create comparison bar charts visualizing model performance metrics across domains."""
    # Define datasets for labels
    datasets = [f"Training\n({train_city})", f"Testing\n({test_city})", f"Validation\n({val_city})"]
    
    # Create figure for regression metrics
    fig, axes = plt.subplots(2, 2, figsize=(14, 10))
    axes = axes.flatten()
    
    # Regression metrics
    regression_metrics = ['MSE', 'MAE', 'RMSE', 'R²']
    metric_values = [
        [train_metrics[m] for m in regression_metrics],
        [test_metrics[m] for m in regression_metrics],
        [val_metrics[m] for m in regression_metrics]
    ]
    
    # Plot regression metrics
    for i, metric in enumerate(regression_metrics):
        values = [metric_values[j][i] for j in range(3)]
        ax = axes[i]
        bars = ax.bar(datasets, values, color=['#1f77b4', '#ff7f0e', '#2ca02c'])
        ax.set_title(f"{metric}", fontsize=12, fontweight='bold')
        ax.set_ylabel('Value')
        
        # Add value labels on top of bars
        for bar in bars:
            height = bar.get_height()
            ax.annotate(f'{height:.4f}',
                       xy=(bar.get_x() + bar.get_width() / 2, height),
                       xytext=(0, 3),  # 3 points vertical offset
                       textcoords="offset points",
                       ha='center', va='bottom', rotation=0, fontsize=9)
    
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, "regression_metrics_comparison.png"), dpi=300, bbox_inches='tight')
    plt.close()
    
    # Create figure for categorical metrics
    cat_metrics = ['Accuracy', 'Hit Rate (POD)', 'False Alarm Rate', 'F1 Score']
    fig, axes = plt.subplots(2, 2, figsize=(14, 10))
    axes = axes.flatten()
    
    cat_values = [
        [train_cat_metrics[m] for m in cat_metrics],
        [test_cat_metrics[m] for m in cat_metrics],
        [val_cat_metrics[m] for m in cat_metrics]
    ]
    
    # Plot categorical metrics
    for i, metric in enumerate(cat_metrics):
        values = [cat_values[j][i] for j in range(3)]
        ax = axes[i]
        bars = ax.bar(datasets, values, color=['#1f77b4', '#ff7f0e', '#2ca02c'])
        ax.set_title(f"{metric}", fontsize=12, fontweight='bold')
        ax.set_ylabel('Value')
        ax.set_ylim(0, 1.0)  # Categorical metrics are usually between 0 and 1
        
        # Add value labels on top of bars
        for bar in bars:
            height = bar.get_height()
            ax.annotate(f'{height:.4f}',
                       xy=(bar.get_x() + bar.get_width() / 2, height),
                       xytext=(0, 3),  # 3 points vertical offset
                       textcoords="offset points",
                       ha='center', va='bottom', rotation=0, fontsize=9)
    
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, "categorical_metrics_comparison.png"), dpi=300, bbox_inches='tight')
    plt.close()
    
    # Create figure for contingency table
    contingency_metrics = ['Hits', 'Misses', 'False Alarms', 'Correct Negatives']
    
    # Calculate percentages for each category
    train_total = sum(train_cat_metrics['Contingency Table'].values())
    test_total = sum(test_cat_metrics['Contingency Table'].values())
    val_total = sum(val_cat_metrics['Contingency Table'].values())
    
    train_pct = [train_cat_metrics['Contingency Table'][m] / train_total * 100 for m in contingency_metrics]
    test_pct = [test_cat_metrics['Contingency Table'][m] / test_total * 100 for m in contingency_metrics]
    val_pct = [val_cat_metrics['Contingency Table'][m] / val_total * 100 for m in contingency_metrics]
    
    # Create a stacked bar chart for percentages
    fig, ax = plt.subplots(figsize=(12, 8))
    
    bar_width = 0.25
    index = np.arange(len(datasets))
    
    # Create color scheme matching the comparison plot colors
    colors = ['blue', 'green', 'red', 'gray']
    
    # Initialize the bottom of each stack
    bottom_train = 0
    bottom_test = 0
    bottom_val = 0
    
    # Add each contingency category as a segment of the stacked bar
    for i, (metric, color) in enumerate(zip(contingency_metrics, colors)):
        # Get values for this metric
        train_val = train_pct[i]
        test_val = test_pct[i]
        val_val = val_pct[i]
        
        # Plot bars for each dataset
        plt.bar(index[0], train_val, bar_width, bottom=bottom_train, color=color, 
               label=metric if i == 0 else "")
        plt.bar(index[1], test_val, bar_width, bottom=bottom_test, color=color)
        plt.bar(index[2], val_val, bar_width, bottom=bottom_val, color=color)
        
        # Add percentage labels inside or above bars
        if train_val > 5:  # Only add label if segment is large enough
            plt.text(index[0], bottom_train + train_val/2, f'{train_val:.1f}%', 
                    ha='center', va='center', color='white', fontweight='bold')
        
        if test_val > 5:
            plt.text(index[1], bottom_test + test_val/2, f'{test_val:.1f}%', 
                    ha='center', va='center', color='white', fontweight='bold')
            
        if val_val > 5:
            plt.text(index[2], bottom_val/2 + val_val/2, f'{val_val:.1f}%', 
                    ha='center', va='center', color='white', fontweight='bold')
        
        # Update the bottom positions for the next set of bars
        bottom_train += train_val
        bottom_test += test_val
        bottom_val += val_val
    
    # Set labels and title
    plt.xlabel('Dataset')
    plt.ylabel('Percentage (%)')
    plt.title('Contingency Table Breakdown (Percentages)', fontsize=14, fontweight='bold')
    plt.xticks(index, datasets)
    plt.ylim(0, 100)
    plt.legend(loc='upper right')
    
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, "contingency_breakdown_comparison.png"), dpi=300, bbox_inches='tight')
    plt.close()

# Function to save metrics to Excel
def save_metrics_to_excel(train_metrics, test_metrics, val_metrics, 
                        train_cat_metrics, test_cat_metrics, val_cat_metrics,
                        train_city, test_city, val_city, output_dir):
    """
    Save all metrics to an Excel file with different tabs for each metric type.
    
    Args:
        train_metrics: Regression metrics for training set
        test_metrics: Regression metrics for testing set
        val_metrics: Regression metrics for validation set
        train_cat_metrics: Categorical metrics for training set
        test_cat_metrics: Categorical metrics for testing set
        val_cat_metrics: Categorical metrics for validation set
        train_city: Training city ID
        test_city: Testing city ID
        val_city: Validation city ID
        output_dir: Directory to save the Excel file
    """
    # Create timestamp for filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = os.path.join(output_dir, f"flood_model_metrics_{timestamp}.xlsx")
    
    # Create a Pandas Excel writer
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # 1. Domain-specific metrics
        domain_mapping = {
            train_city: {"type": "Training", "reg": train_metrics, "cat": train_cat_metrics},
            test_city: {"type": "Testing", "reg": test_metrics, "cat": test_cat_metrics},
            val_city: {"type": "Validation", "reg": val_metrics, "cat": val_cat_metrics}
        }
        
        for domain, metrics in domain_mapping.items():
            # Create DataFrame for regression metrics
            reg_metrics = pd.DataFrame([{
                'Metric': metric,
                'Value': metrics['reg'][metric]
            } for metric in ['MSE', 'MAE', 'RMSE', 'R²']])
            
            # Create DataFrame for categorical metrics
            cat_metrics_list = ['Accuracy', 'Hit Rate (POD)', 'False Alarm Rate', 
                             'Critical Success Index', 'Precision', 'F1 Score', 'Bias Score']
            cat_metrics = pd.DataFrame([{
                'Metric': metric,
                'Value': metrics['cat'][metric]
            } for metric in cat_metrics_list if metric in metrics['cat']])
            
            # Create DataFrame for contingency table
            cont_metrics = pd.DataFrame([{
                'Category': metric,
                'Count': metrics['cat']['Contingency Table'][metric],
                'Percentage': metrics['cat']['Contingency Table'][metric] / sum(metrics['cat']['Contingency Table'].values()) * 100
            } for metric in ['Hits', 'Misses', 'False Alarms', 'Correct Negatives']])
            
            # Write to Excel
            sheet_name = f"{domain}_{metrics['type']}"
            reg_metrics.to_excel(writer, sheet_name=sheet_name, startrow=1, startcol=0, index=False)
            cat_metrics.to_excel(writer, sheet_name=sheet_name, startrow=len(reg_metrics) + 4, startcol=0, index=False)
            cont_metrics.to_excel(writer, sheet_name=sheet_name, startrow=len(reg_metrics) + len(cat_metrics) + 7, startcol=0, index=False)
            
            # Add headers
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            worksheet.cell(row=1, column=1, value=f"{domain} - {metrics['type']} Metrics").font = Font(bold=True, size=14)
            worksheet.cell(row=2, column=1, value="Regression Metrics").font = Font(bold=True)
            worksheet.cell(row=len(reg_metrics) + 5, column=1, value="Categorical Metrics").font = Font(bold=True)
            worksheet.cell(row=len(reg_metrics) + len(cat_metrics) + 8, column=1, value="Contingency Table").font = Font(bold=True)
        
        # 2. Cross-domain comparison - Regression Metrics
        regression_metrics = ['MSE', 'MAE', 'RMSE', 'R²']
        reg_comparison = pd.DataFrame([{
            'Metric': metric,
            f'{train_city} (Training)': train_metrics[metric],
            f'{test_city} (Testing)': test_metrics[metric],
            f'{val_city} (Validation)': val_metrics[metric]
        } for metric in regression_metrics])
        
        reg_comparison.to_excel(writer, sheet_name='Regression_Comparison', startrow=1, index=False)
        worksheet = writer.sheets['Regression_Comparison']
        worksheet.cell(row=1, column=1, value="Regression Metrics Comparison Across Domains").font = Font(bold=True, size=14)
        
        # 3. Cross-domain comparison - Categorical Metrics
        categorical_metrics = ['Accuracy', 'Hit Rate (POD)', 'False Alarm Rate', 'F1 Score', 'Precision']
        cat_comparison = pd.DataFrame([{
            'Metric': metric,
            f'{train_city} (Training)': train_cat_metrics[metric],
            f'{test_city} (Testing)': test_cat_metrics[metric],
            f'{val_city} (Validation)': val_cat_metrics[metric]
        } for metric in categorical_metrics])
        
        cat_comparison.to_excel(writer, sheet_name='Categorical_Comparison', startrow=1, index=False)
        worksheet = writer.sheets['Categorical_Comparison']
        worksheet.cell(row=1, column=1, value="Categorical Metrics Comparison Across Domains").font = Font(bold=True, size=14)
        
        # 4. Cross-domain comparison - Contingency Table
        contingency_metrics = ['Hits', 'Misses', 'False Alarms', 'Correct Negatives']
        
        # Counts
        cont_comparison_counts = pd.DataFrame([{
            'Category': metric,
            f'{train_city} (Training)': train_cat_metrics['Contingency Table'][metric],
            f'{test_city} (Testing)': test_cat_metrics['Contingency Table'][metric],
            f'{val_city} (Validation)': val_cat_metrics['Contingency Table'][metric]
        } for metric in contingency_metrics])
        
        # Percentages
        train_total = sum(train_cat_metrics['Contingency Table'].values())
        test_total = sum(test_cat_metrics['Contingency Table'].values())
        val_total = sum(val_cat_metrics['Contingency Table'].values())
        
        cont_comparison_pct = pd.DataFrame([{
            'Category': metric,
            f'{train_city} (Training) %': train_cat_metrics['Contingency Table'][metric] / train_total * 100,
            f'{test_city} (Testing) %': test_cat_metrics['Contingency Table'][metric] / test_total * 100,
            f'{val_city} (Validation) %': val_cat_metrics['Contingency Table'][metric] / val_total * 100
        } for metric in contingency_metrics])
        
        cont_comparison_counts.to_excel(writer, sheet_name='Contingency_Comparison', startrow=1, index=False)
        cont_comparison_pct.to_excel(writer, sheet_name='Contingency_Comparison', startrow=len(cont_comparison_counts) + 4, index=False)
        
        worksheet = writer.sheets['Contingency_Comparison']
        worksheet.cell(row=1, column=1, value="Contingency Table Comparison Across Domains").font = Font(bold=True, size=14)
        worksheet.cell(row=2, column=1, value="Counts").font = Font(bold=True)
        worksheet.cell(row=len(cont_comparison_counts) + 5, column=1, value="Percentages (%)").font = Font(bold=True)
    
    print(f"\nMetrics saved to Excel file: {excel_path}")
    return excel_path

# Main execution
if __name__ == "__main__":
    # Load pre-trained model
    checkpoint_path = '/home/users/li1995/global_flood/FloodRisk-DL/src/checkpoints/epoch=36-val_loss=1.2045.ckpt'

    # Create model instance
    model = FloodPredictionModel.load_from_checkpoint(checkpoint_path=checkpoint_path, config=config, strict=False)

    # Set model to evaluation mode
    model.to('cuda')
    model.eval()

    # Load city rainfall data
    with open('../src/cities_rainfall.json', 'r') as f:
        cities_data = json.load(f)
    
    # Process training data (HOU002)
    train_city = "HOU002"
    train_rainfall = None
    for city in cities_data:
        if city['City ID'] == train_city:
            train_rainfall = city['100-yr']
            break

    if train_rainfall:
        train_dem, train_target, train_pred, train_meta, train_cat_metrics = process_city(train_city, train_rainfall, model)

    # Process testing data (HOU007)
    test_city = "HOU007"
    test_rainfall = None
    for city in cities_data:
        if city['City ID'] == test_city:
            test_rainfall = city['100-yr']
            break

    if test_rainfall:
        test_dem, test_target, test_pred, test_meta, test_cat_metrics = process_city(test_city, test_rainfall, model)

    # Process validation data (DAL002)
    val_city = "DAL002"
    val_rainfall = None
    for city in cities_data:
        if city['City ID'] == val_city:
            val_rainfall = city['100-yr']
            break

    if val_rainfall:
        val_dem, val_target, val_pred, val_meta, val_cat_metrics = process_city(val_city, val_rainfall, model)
    
    # Calculate and display metrics for all datasets
    train_metrics = calculate_metrics(train_target, train_pred)
    test_metrics = calculate_metrics(test_target, test_pred)
    val_metrics = calculate_metrics(val_target, val_pred)

    print("\nRegression Metrics:")
    print(f"Training Data ({train_city}):")
    for metric, value in train_metrics.items():
        print(f"  {metric}: {value:.4f}")

    print(f"\nTesting Data ({test_city}):")
    for metric, value in test_metrics.items():
        print(f"  {metric}: {value:.4f}")

    print(f"\nValidation Data ({val_city}):")
    for metric, value in val_metrics.items():
        print(f"  {metric}: {value:.4f}")
    
    # Create comparison plot
    comparison_output_path = os.path.join(OUTPUT_DIR, "flood_prediction_comparison.png")
    create_comparison_plot(
        train_data=(train_city, train_dem, train_target, train_pred, train_meta),
        test_data=(test_city, test_dem, test_target, test_pred, test_meta),
        val_data=(val_city, val_dem, val_target, val_pred, val_meta),
        output_path=comparison_output_path
    )
    
    # Create metrics visualizations
    create_metrics_visualizations(
        train_metrics, test_metrics, val_metrics,
        train_cat_metrics, test_cat_metrics, val_cat_metrics,
        train_city, test_city, val_city, OUTPUT_DIR
    )
    
    # Create comprehensive summary tables
    print("\n" + "="*100)
    print("                                COMPREHENSIVE MODEL EVALUATION SUMMARY                                ")
    print("="*100)
    
    # Define the metrics to display
    regression_metrics = ['MSE', 'MAE', 'RMSE', 'R²']
    categorical_metrics = [
        'Accuracy', 
        'Hit Rate (POD)', 
        'False Alarm Rate', 
        'Critical Success Index',
        'Precision',
        'F1 Score'
    ]
    contingency_metrics = ['Hits', 'Misses', 'False Alarms', 'Correct Negatives']
    
    # Define domain-data type mapping (which city is used for which data type)
    domain_mapping = {
        train_city: "Training",
        test_city: "Testing",
        val_city: "Validation"
    }
    
    # Create metrics mapping
    metrics_data = {
        "Training": {"reg": train_metrics, "cat": train_cat_metrics},
        "Testing": {"reg": test_metrics, "cat": test_cat_metrics},
        "Validation": {"reg": val_metrics, "cat": val_cat_metrics}
    }
    
    # Display metrics for each domain
    for domain, data_type in domain_mapping.items():
        metrics = metrics_data[data_type]
        
        print(f"\n\n--- Domain: {domain} ({data_type} Data) ---")
        
        # Regression metrics
        print("\nRegression Metrics:")
        print(f"{'Metric':<25} {'Value':>15}")
        print("-" * 42)
        for metric in regression_metrics:
            print(f"{metric:<25} {metrics['reg'][metric]:>15.5f}")
        
        # Categorical metrics
        print("\nCategorical Metrics:")
        print(f"{'Metric':<25} {'Value':>15}")
        print("-" * 42)
        for metric in categorical_metrics:
            print(f"{metric:<25} {metrics['cat'][metric]:>15.5f}")
        
        # Contingency table
        print("\nContingency Table:")
        print(f"{'Category':<25} {'Count':>15} {'Percentage':>15}")
        print("-" * 57)
        
        # Calculate total pixels
        counts = metrics['cat']['Contingency Table']
        total = sum(counts.values())
        
        # Print contingency table with percentages
        for metric in contingency_metrics:
            count = counts[metric]
            percentage = (count / total) * 100 if total > 0 else 0
            print(f"{metric:<25} {count:>15} {percentage:>14.2f}%")
    
    # Create a cross-domain comparison table for key metrics
    print("\n\n" + "="*100)
    print("                                CROSS-DOMAIN COMPARISON SUMMARY                                ")
    print("="*100)
    
    # Format header for cross-domain table
    header = f"{'Metric':<25} "
    for domain in [train_city, test_city, val_city]:
        data_type = domain_mapping[domain]
        formatted_domain = f"{domain} ({data_type})"
        header += f"{formatted_domain:>20} "
    
    print(header)
    print("-" * len(header))
    
    # Print regression metrics across domains
    print("\nRegression Metrics:")
    for metric in regression_metrics:
        row = f"{metric:<25} "
        for domain, data_type in domain_mapping.items():
            row += f"{metrics_data[data_type]['reg'][metric]:>20.5f} "
        print(row)
    
    # Print key categorical metrics across domains
    print("\nCategorical Metrics:")
    selected_cat_metrics = ['Accuracy', 'Hit Rate (POD)', 'F1 Score']
    for metric in selected_cat_metrics:
        row = f"{metric:<25} "
        for domain, data_type in domain_mapping.items():
            row += f"{metrics_data[data_type]['cat'][metric]:>20.5f} "
        print(row)
    
    # Create comparison visualizations
    print("\n" + "="*80)
    print("             VISUALIZATIONS HAVE BEEN CREATED IN THE OUTPUT DIRECTORY              ")
    print("="*80)
    print("\nFiles generated:")
    print("\n1. Individual Domain Prediction Maps:")
    print(f"   - {train_city}_{train_rainfall}_prediction.png")
    print(f"   - {test_city}_{test_rainfall}_prediction.png")
    print(f"   - {val_city}_{val_rainfall}_prediction.png")
    
    print("\n2. Domain-Specific Comprehensive Metrics:")
    print(f"   - {train_city}_metrics.png")
    print(f"   - {test_city}_metrics.png")
    print(f"   - {val_city}_metrics.png")
    
    print("\n3. Cross-Domain Comparison Visualizations:")
    print(f"   - flood_prediction_comparison.png (Side-by-side map comparison)")
    print(f"   - regression_metrics_comparison.png (MSE, MAE, RMSE, R² comparison)")
    print(f"   - categorical_metrics_comparison.png (Classification metrics comparison)")
    print(f"   - contingency_breakdown_comparison.png (Contingency table comparison)")
    
    print(f"\nOutput directory: {os.path.abspath(OUTPUT_DIR)}")

    # Save metrics to Excel
    save_metrics_to_excel(
        train_metrics, test_metrics, val_metrics,
        train_cat_metrics, test_cat_metrics, val_cat_metrics,
        train_city, test_city, val_city, OUTPUT_DIR
    )
